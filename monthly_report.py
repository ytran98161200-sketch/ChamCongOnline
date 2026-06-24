
import calendar
import pandas as pd
from io import BytesIO
from sqlalchemy import text
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from database import engine


def get_monthly_report(month, year):

    with engine.connect() as conn:

        employees = conn.execute(
            text("""
                SELECT
                    employee_code,
                    fullname,
                    department
                FROM employees
                ORDER BY department, fullname
            """)
        ).fetchall()

        logs = conn.execute(
            text("""
                SELECT
                    employee_code,
                    DATE(scan_time) AS work_date,
                    COUNT(*) AS scan_count
                FROM attendance_logs
                WHERE
                    EXTRACT(MONTH FROM scan_time)=:month
                    AND EXTRACT(YEAR FROM scan_time)=:year
                GROUP BY
                    employee_code,
                    DATE(scan_time)
                ORDER BY
                    employee_code,
                    work_date
            """),
            {
                "month": month,
                "year": year
            }
        ).fetchall()

    attendance = {}

    for employee_code, work_date, scan_count in logs:
        attendance[
            (employee_code, work_date.day)
        ] = scan_count

    days_in_month = calendar.monthrange(
        year,
        month
    )[1]

    report = []

    for employee_code, fullname, department in employees:

        row = [
            employee_code,
            fullname,
            department
        ]

        total = 0

        for day in range(
            1,
            days_in_month + 1
        ):

            count = attendance.get(
                (employee_code, day),
                0
            )

            if count >= 4:
                value = 1
                total += 1

            elif count >= 2:
                value = 0.5
                total += 0.5

            else:
                value = ""

            row.append(value)

        row.append(total)
        row.append(26)
        row.append(0)
        row.append(0)
        row.append(total)
        row.append(total * 8)
        row.append(0)
        row.append(0)
        row.append(0)

        report.append(row)

    columns = [
        "Mã NV",
        "Họ tên",
        "Phòng ban"
    ]

    for day in range(
        1,
        days_in_month + 1
    ):
        columns.append(str(day))

    columns.extend([
        "Tổng công",
        "Công chuẩn",
        "Nghỉ phép",
        "Nghỉ không lương",
        "Ngày trả lương",
        "Giờ trả lương",
        "Đi trễ",
        "Về sớm",
        "Tăng ca"
    ])

    return pd.DataFrame(
        report,
        columns=columns
    )


def export_monthly_report(month, year):

    df = get_monthly_report(
        month,
        year
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="BCCVP",
            index=False,
            startrow=2
        )
        last_data_row = len(df) + 3
        ws = writer.sheets["BCCVP"]

        from openpyxl.utils import get_column_letter

        total_cols = len(df.columns)
        last_col = get_column_letter(total_cols)

        ws.merge_cells(f"A1:{last_col}1")

        ws["A1"] = (
            f"BÁO CÁO CHẤM CÔNG THÁNG {month}/{year}"
        )

        ws["A1"].font = Font(
            size=18,
            bold=True
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 35
        blue_fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for cell in ws[3]:
            cell.fill = blue_fill
            cell.font = Font(
                color="FFFFFF",
                bold=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = thin_border

        days_in_month = calendar.monthrange(
            year,
            month
        )[1]

        for day in range(
            1,
            days_in_month + 1
        ):

            col = 4 + day

            weekday = calendar.weekday(
                year,
                month,
                day
            )

            thu = {
                0: "T2",
                1: "T3",
                2: "T4",
                3: "T5",
                4: "T6",
                5: "T7",
                6: "CN"
            }[weekday]

            ws.cell(
                2,
                col
            ).value = thu

            ws.cell(
                2,
                col
            ).alignment = Alignment(
                horizontal="center"
            )

            if weekday == 6:
                for r in range(4, 201):
                    ws.cell(
                        r,
                        col
                    ).fill = PatternFill(
                        "solid",
                        fgColor="F2F2F2"
                    )
                ws.cell(
                    2,
                    col
                ).fill = PatternFill(
                    "solid",
                    fgColor="FF0000"
                )

                ws.cell(
                    2,
                    col
                ).font = Font(
                    color="FFFFFF",
                    bold=True
                )
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
    for r in range(
        5,
        ws.max_row + 1
    ):
        ws.cell(
            r,
            2
        ).alignment = Alignment(
            horizontal="left",
            vertical="center"
        )
    from openpyxl.utils import get_column_letter

    for col in range(
        1,
        ws.max_column + 1
    ):
        letter = get_column_letter(col)

        if col == 1:
            ws.column_dimensions[letter].width = 12

        elif col == 2:
            ws.column_dimensions[letter].width = 30

        elif col == 3:
            ws.column_dimensions[letter].width = 20

        elif col <= days_in_month + 3:
            ws.column_dimensions[letter].width = 7

        else:
            ws.column_dimensions[letter].width = 14
    ws.freeze_panes = "D4"
    output.seek(0)

    return output

